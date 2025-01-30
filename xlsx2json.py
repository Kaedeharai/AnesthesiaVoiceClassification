import openpyxl
import json
import io


def excel_to_json(excel_file, json_file_name):
    book = openpyxl.load_workbook(excel_file)
    sheet = book["Sheet1"]
    max_row = sheet.max_row
    max_column = sheet.max_column

    result = []
    heads = []

    for column in range(max_column):
        heads.append(sheet.cell(1, column + 1).value)
    for row in range(max_row):
        if row == 0:
            continue
        one_line = {}

        for column in range(max_column):
            k = heads[column]
            cell = sheet.cell(row + 1, column + 1)
            value = cell.value
            one_line[k] = value
        print(one_line)
        result.append(one_line)
    book.close()

    save_json_file(result, json_file_name)


def save_json_file(jd, json_file_name):
    file = io.open(json_file_name, 'w', encoding='utf-8')
    txt = json.dumps(jd, indent=2, ensure_ascii=False)
    file.write(txt)
    file.close()
    
if '__main__' == __name__:
     excel_to_json(u'长庚医院AI语言数据.xlsx', 'before.json')
